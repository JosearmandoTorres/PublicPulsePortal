from fastapi import FastAPI, UploadFile, File, HTTPException, status
from fastapi.middleware.cors import CORSMiddleware
from app.db_selections import (
    add_selection as db_add_selection,
    remove_selection as db_remove_selection,
    list_selections as db_list_selections,
)

import os
import sqlite3
import uuid
from datetime import datetime
import csv
from typing import List, Dict

app = FastAPI()

# -------------------------------------------------------------------
# Config
# -------------------------------------------------------------------
DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "data")
DB_PATH = os.path.join(DATA_DIR, "ppp.db")
UPLOADS_DIR = os.path.join(DATA_DIR, "uploads")

# -------------------------------------------------------------------
# DB init
# -------------------------------------------------------------------
def init_db():
    os.makedirs(DATA_DIR, exist_ok=True)
    os.makedirs(UPLOADS_DIR, exist_ok=True)
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()

        # datasets registry
        c.execute("""
            CREATE TABLE IF NOT EXISTS datasets (
                id TEXT PRIMARY KEY,
                filename TEXT,
                stored_path TEXT,
                uploaded_at TEXT
            )
        """)

        # raw_rows for parsed questions/responses
        c.execute("""
            CREATE TABLE IF NOT EXISTS raw_rows (
                row_id INTEGER PRIMARY KEY AUTOINCREMENT,
                dataset_id TEXT,
                QuestionID TEXT,
                RespTxt TEXT,
                RespPct TEXT,
                QuestionTxt TEXT,
                QuestionNote TEXT,
                SubPopulation TEXT,
                ReleaseDate TEXT,
                SurveyOrg TEXT,
                SurveySponsor TEXT,
                SourceDoc TEXT,
                BegDate TEXT,
                EndDate TEXT,
                Country TEXT,
                SampleDesc TEXT,
                SampleSize TEXT,
                IntMethod TEXT,
                StudyNote TEXT,
                Topics TEXT,
                SampleTypes TEXT,
                DatePublished TEXT,
                Link TEXT
            )
        """)

        # selections: per-user cart of question_ids in a dataset
        c.execute("""
            CREATE TABLE IF NOT EXISTS selections (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id TEXT NOT NULL,
                dataset_id TEXT NOT NULL,
                question_id TEXT NOT NULL,
                created_at TEXT NOT NULL,
                UNIQUE(user_id, dataset_id, question_id)
            )
        """)

        conn.commit()

init_db()

# -------------------------------------------------------------------
# Middleware
# -------------------------------------------------------------------
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://127.0.0.1:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# -------------------------------------------------------------------
# Health
# -------------------------------------------------------------------
@app.get("/")
def root():
    return {"message": "Welcome to PublicPulsePortal backend 🚀"}

@app.get("/health")
def health():
    return {"status": "ok"}

# -------------------------------------------------------------------
# Helpers for ingest
# -------------------------------------------------------------------
REQUIRED_HEADERS = [
    "QuestionID", "RespTxt", "RespPct", "QuestionTxt",
    "ReleaseDate", "SurveyOrg", "Country", "SampleSize", "SampleDesc", "Link"
]

RAW_COLS: List[str] = [
    "QuestionID", "RespTxt", "RespPct", "QuestionTxt", "QuestionNote",
    "SubPopulation", "ReleaseDate", "SurveyOrg", "SurveySponsor", "SourceDoc",
    "BegDate", "EndDate", "Country", "SampleDesc", "SampleSize", "IntMethod",
    "StudyNote", "Topics", "SampleTypes", "DatePublished", "Link"
]

def _validate_headers(found: List[str]) -> List[str]:
    return [h for h in REQUIRED_HEADERS if h not in found]

def _rows_from_csv(path: str) -> List[Dict[str, str]]:
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        missing = _validate_headers(headers)
        if missing:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={"error": "Missing required headers", "missing": missing}
            )
        rows = []
        for r in reader:
            rows.append({col: (r.get(col) or "") for col in RAW_COLS})
        return rows

def _rows_from_xlsx(path: str) -> List[Dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"error": "XLSX support requires 'openpyxl'. Please add to requirements.txt"}
        )
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    header = [str(c.value).strip() if c.value is not None else "" 
              for c in next(ws.iter_rows(min_row=1, max_row=1))]
    missing = _validate_headers(header)
    if missing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"error": "Missing required headers", "missing": missing}
        )
    idx = {name: i for i, name in enumerate(header)}
    rows: List[Dict[str, str]] = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        rec: Dict[str, str] = {}
        for col in RAW_COLS:
            if col in idx:
                cell = row[idx[col]]
                val = "" if cell.value is None else str(cell.value)
                rec[col] = val
            else:
                rec[col] = ""
        rows.append(rec)
    return rows

def _ingest_rows(dataset_id: str, rows: List[Dict[str, str]]) -> int:
    if not rows:
        return 0
    placeholders = ",".join(["?"] * (1 + len(RAW_COLS)))
    cols_sql = ",".join(["dataset_id"] + RAW_COLS)
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.executemany(
            f"INSERT INTO raw_rows ({cols_sql}) VALUES ({placeholders})",
            [
                tuple([dataset_id] + [r.get(col, "") for col in RAW_COLS])
                for r in rows
            ],
        )
        conn.commit()
        return c.rowcount

# -------------------------------------------------------------------
# Endpoints
# -------------------------------------------------------------------
@app.post("/datasets/upload")
async def upload_dataset(file: UploadFile = File(...)):
    dataset_id = str(uuid.uuid4())
    original_name = file.filename
    stored_name = f"{dataset_id}_{original_name}"
    stored_path = os.path.join(UPLOADS_DIR, stored_name)

    # Save file
    with open(stored_path, "wb") as f:
        content = await file.read()
        f.write(content)

    # Register in datasets
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute(
            "INSERT INTO datasets (id, filename, stored_path, uploaded_at) VALUES (?, ?, ?, ?)",
            (dataset_id, original_name, stored_path, datetime.utcnow().isoformat())
        )
        conn.commit()

    # Ingest into raw_rows
    ext = os.path.splitext(original_name)[1].lower()
    try:
        if ext == ".csv":
            rows = _rows_from_csv(stored_path)
        elif ext in (".xlsx", ".xlsm"):
            rows = _rows_from_xlsx(stored_path)
        else:
            return {
                "ok": True,
                "dataset_id": dataset_id,
                "filename": original_name,
                "stored_path": stored_path,
                "rows_ingested": 0,
                "note": "Unsupported ingest format; only .csv, .xlsx, .xlsm are parsed."
            }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail={"error": "Ingest failed", "message": str(e)})

    ingested = _ingest_rows(dataset_id, rows)

    return {
        "ok": True,
        "dataset_id": dataset_id,
        "filename": original_name,
        "stored_path": stored_path,
        "rows_ingested": ingested
    }

@app.get("/datasets")
def list_datasets(limit: int = 50, offset: int = 0):
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute("SELECT COUNT(*) FROM datasets")
        total = c.fetchone()[0]
        c.execute(
            "SELECT id, filename, stored_path, uploaded_at FROM datasets ORDER BY uploaded_at DESC LIMIT ? OFFSET ?",
            (limit, offset)
        )
        items = [
            {"id": r[0], "filename": r[1], "stored_path": r[2], "uploaded_at": r[3]}
            for r in c.fetchall()
        ]
    return {"total": total, "items": items}

# -------------------------------------------------------------------
# Blocks API (dataset_id optional = global database view)
# -------------------------------------------------------------------
from typing import Optional

@app.get("/questions/blocks")
def get_question_blocks(
    dataset_id: Optional[str] = None,
    search: Optional[str] = None,
    limit: int = 25,
    offset: int = 0,
):
    """
    Returns question 'blocks'. If dataset_id is provided, scope to that dataset.
    If omitted, return across ALL datasets.

    Each block groups rows by (dataset_id, full QuestionID), and includes dataset_id
    so the frontend can display source info.
    """
    params = []
    where = "WHERE 1=1"
    if dataset_id:
        where += " AND dataset_id = ?"
        params.append(dataset_id)
    if search:
        where += " AND QuestionTxt LIKE ?"
        params.append(f"%{search}%")

    sql = f"""
        SELECT
            dataset_id,
            QuestionID,
            RespTxt, RespPct,
            QuestionTxt,
            ReleaseDate, SurveyOrg, SurveySponsor, Country,
            SampleSize, SampleDesc, Link,
            row_id
        FROM raw_rows
        {where}
        ORDER BY dataset_id, QuestionID, row_id
    """

    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        rows = c.execute(sql, params).fetchall()

    # Group by (dataset_id, QuestionID)
    blocks = {}
    order_keys = []
    for (
        ds_id, qid, resp_txt, resp_pct, qtxt,
        release_date, survey_org, survey_sponsor, country,
        sample_size, sample_desc, link, row_id
    ) in rows:
        qid_str = str(qid) if qid is not None else ""
        key = (ds_id or "", qid_str)
        if key not in blocks:
            blocks[key] = {
                "dataset_id": ds_id,
                "question_id": qid_str,
                "question_text": qtxt or "",
                "metadata": {
                    "ReleaseDate": release_date or "",
                    "SurveyOrg": survey_org or "",
                    "SurveySponsor": survey_sponsor or "",
                    "Country": country or "",
                    "SampleSize": sample_size or "",
                    "SampleDesc": sample_desc or "",
                    "Link": link or "",
                },
                "responses": []
            }
            order_keys.append(key)
        blocks[key]["responses"].append({
            "RespTxt": resp_txt or "",
            "value": resp_pct or ""
        })

    total = len(order_keys)
    page_keys = order_keys[offset: offset + limit]
    items = [blocks[k] for k in page_keys]
    return {"total": total, "items": items}


# -------------------------------------------------------------------
# Selections API (Phase 4)
# -------------------------------------------------------------------
from pydantic import BaseModel

class SelectionIn(BaseModel):
    user_id: str
    dataset_id: str
    question_id: str

@app.post("/selections")
def add_selection(sel: SelectionIn):
    try:
        with sqlite3.connect(DB_PATH) as conn:
            db_add_selection(conn, sel.user_id, sel.dataset_id, sel.question_id)
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=400, detail={"error": "insert_failed", "message": str(e)})


@app.delete("/selections")
def remove_selection(user_id: str, dataset_id: str, question_id: str):
    with sqlite3.connect(DB_PATH) as conn:
        db_remove_selection(conn, user_id, dataset_id, question_id)
    return {"ok": True}


@app.get("/selections")
def list_selections(user_id: str, dataset_id: str):
    with sqlite3.connect(DB_PATH) as conn:
        items = db_list_selections(conn, user_id, dataset_id)
    return {"total": len(items), "items": items}
